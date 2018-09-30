#!/usr/bin/env python3

# -------------------------------------------------------------------------------
# ------------------------                               ------------------------
# ------------------------      MD5, SHA1,and SHA256     ------------------------
# ------------------------                               ------------------------
# -------------------------------------------------------------------------------

import hashlib  # md5/sha1/sha256
import os  # traverse directories
import time
import platform  # detect o/s
import re  # regular expression
import sys  # command line args
import openpyxl  # excel sheets
from pathlib import Path  # deals with paths between operating systems
from pathlib import PureWindowsPath
import threading  #used to have a separate thread count total files


#main will mostly act as a menu system, the core functionality goes to functions
def main():

    #skip menu if user provides valid argument, if not spit out an example
    if len(sys.argv) > 1 and len(
            sys.argv
    ) <= 4:  #if user passes hundreds of arguments that is there own fault
        validArgs()
        sys.exit(0)
    elif len(
            sys.argv
    ) >= 6:  #not going to incorporate single file mode for commandline this version
        option = validInput(sys.argv[1], 1, 4)
        path = getPath("", sys.argv[2])  #check dir
        t1 = threading.Thread(
            target=fileCount, args=(path,))  #start total file count immediately
        t1.start()
        #if thread is running let user know we are calculating files
        if t1.isAlive():
            print("Calculating total files...")
        output = getPath("", sys.argv[3])  #check dir
        filename = sys.argv[4] + ".xlsx"  #filename
        recursive = recursCheck(sys.argv[5])
        if recursive:  #if valid check true/false
            print("Recursion enabled\n")
        else:
            print("Recursion failed wrong argument passed\n")
        t1.join()  #join thread before continueing
        fileHash(option, path, output, filename, recursive)
    elif len(sys.argv) == 5:  #no recursion
        option = validInput(sys.argv[1], 1, 4)
        path = getPath("", sys.argv[2])  #check dir
        output = getPath("", sys.argv[3])  #check dir
        filename = sys.argv[4] + ".xlsx"  #filename
        fileHash(option, path, output, filename)
    else:  #menu
        #print out a menu and guide user, if arguments are passed we can bypass this
        print("", 60 * '-', "\n Hash sheet generator |v-1.0|", "\n", 60 * '-',
              "\n 1. MD5, SHA-1, & SHA-256", "\n 2. MD5", "\n 3. SHA-1",
              "\n 4. SHA-256", "\n 5. Single file mode", "\n 6. Help",
              "\n 0. Quit", "\n", 60 * '-')
        option = input(" Option: ")

        #get options to execute calculations
        option = validInput(option, 1, 5)  #user input, start range, end range
        if option == "5":
            path = getPath("Directory of file: ")  #starting directory
            single = getFile()  #get file name
            singleMode(path, single)
            sys.exit(0)  # we will exit here
        else:
            path = getPath("Directory to hash: ")  #starting directory

            t1 = threading.Thread(
                target=fileCount,
                args=(path,))  #start total file count immediately
            t1.start()
        output = getPath("Directory to save file: ")  #save directory
        filename = input("File name: ")  #any name they wish not our problem
        filename += ".xlsx"

        #ask if this should be recursive before we start
        recursive = recursCheck()

        #if thread is running let user know we are calculating files
        if t1.isAlive():
            print("Calculating total files...")

        t1.join()  #make sure this ends before going further

        #execute
        fileHash(option, path, output, filename, recursive)
        #close program
        sys.exit(0)  #success


# ------------------------------------------------------------------------------
# FUNCTIONS
# ------------------------------------------------------------------------------
#get total file count we will be processing
def fileCount(path):
    global totalFiles
    for root, dirs, files in os.walk(path):
        totalFiles += len(files)


#set recursive true or none
def recursCheck(recursive=None):
    #quick dictionary for valid inputs down here
    traverse = {  #swap to generator function
            "yes": True,
            "y": True,
            "Y": True,
            "no": False,
            "n": False,
            "N": False,
            "r": True,
            "R": True,
            "1": True,
            "0": False,
    }

    if recursive == None:
        #ask if this should be recursive before we start
        while True:  #no exiting on this 1 they can close the window at this point
            recursive = input("Recursive (y/n)?: ")

            if recursive in traverse:  #if valid check true/false
                if traverse[recursive]:
                    return traverse[recursive]
    else:
        if recursive in traverse:  #if valid check true/false
            if traverse[recursive]:
                return traverse[recursive]
            else:
                return False


#a single file mode that only prints to screen
def singleMode(path, single):
    #print message
    print("\n\n", 33 * '-', "\nSingle file mode", "\n", 33 * '-', "\n 1 -> MD5",
          "\n 2 -> SHA-1", "\n 3 -> SHA-256", "\n 0 -> Quit", "\n", 33 * '-')

    option = input(" Option: ")

    while True:
        option = validInput(option, 1, 3)  #user input, start range, end range
        print("")  # add a space

        if option == "1":
            with open((os.path.join(path, single)),
                      'rb') as hashMe:  #read in binary
                contents = hashMe.read()  #read in the file
                print("MD5: ", md5File(contents))
        elif option == "2":
            with open((os.path.join(path, single)),
                      'rb') as hashMe:  #read in binary
                contents = hashMe.read()  #read in the file
                print("SHA-1: ", sh1File(contents))
        elif option == "3":
            with open((os.path.join(path, single)),
                      'rb') as hashMe:  #read in binary
                contents = hashMe.read()  #read in the file
                print("SHA-256: ", sh56File(contents))
        else:
            print("Invalid choice")

        option = input(
            "\n Option: ")  #see if user wants other hashes on file or quit


def getFile():
    singleFile = input("Name of file: ")
    while True:  # run loop unless broken out of
        # first we test if the users file exist
        try:
            with open(singleFile, 'r'):  #check file
                return singleFile  #return file name
        except FileNotFoundError:
            singleFile = input(
                "File not found, enter new file name or 0 to quit: "
            )  # Catch and ask user for new file or if they wish to quit
            if singleFile == "0":
                sys.exit(0)


#create new book
def newBook(path):
    hashBook = openpyxl.Workbook()  #create workbook to save spreadsheet
    hashSheet = hashBook.active  #get the first sheet and use filename to modify
    hashSheet.title = os.path.basename(path)  # set initial sheet
    return hashBook


#format sheet header
def sheetHeader(hashBook, option, sheetNum=0):
    sheet = hashBook.sheetnames  #sheet
    hashSheet = hashBook[sheet[sheetNum]]  #setsheet

    hashType = ["File Name", "MD5", "SHA-1",
                "SHA-256"]  # header for spreadsheet
    #set header by options
    if option == "1":
        for col in range(1, 5):  #set cells
            hashSheet.cell(row=1, column=col).value = hashType[col - 1]
        return hashSheet
    if option == "2":
        for col in range(1, 3):
            hashSheet.cell(row=1, column=col).value = hashType[col - 1]
        return hashSheet
    if option == "3":

        hashSheet.cell(row=1, column=1).value = hashType[0]
        hashSheet.cell(row=1, column=2).value = hashType[2]
        return hashSheet
    if option == "4":

        hashSheet.cell(row=1, column=1).value = hashType[0]
        hashSheet.cell(row=1, column=2).value = hashType[3]
        return hashSheet


#take input traverse directory and save
def fileHash(option, path, output, filename, recursive=None):

    hashBook = newBook(path)  #create workbook to save spreadsheet
    hashSheet = sheetHeader(hashBook, option)  #format header set sheet

    sheetNum = 1  #sheet changer
    i = 0
    filesSkipped = []
    #run different loops depending on user options maybe more separate functions for each user option

    #start timer
    start = time.time()
    if option == "1":
        for cwd, subList, file in os.walk(
                path):  #start from selected directory and branch out

            if os.path.basename(cwd) != os.path.basename(path) and len(
                    file) != 0:
                # create new sheet
                sheetName = str(os.path.basename(cwd).encode('utf-8'))
                sheetName = re.sub(r'[^\w]', ' ', sheetName)
                sheetName = '{:1.29}'.format(
                    sheetName)  #max hardcoded sheetname 31 we limit to 30
                hashBook.create_sheet(sheetName)
                hashSheet = sheetHeader(hashBook, option,
                                        sheetNum)  # reprep sheet

                sheetNum += 1

            for name in file:
                i += 1
                print(
                    ">>> Processed: {} out of {} files".format(i, totalFiles),
                    end='\r',
                    flush=True)

                try:
                    with open((os.path.join(cwd, name)),
                              'rb') as hashMe:  #read in binary

                        contents = hashMe.read()  #read in the file
                        hashes = [  #create a list containing file name and digest for each hash
                            name,
                            md5File(contents),
                            sh1File(contents),
                            sh56File(contents)
                        ]
                    hashSheet.append(hashes)

                except IOError:

                    filesSkipped.append(os.path.join(cwd, name))

            if not recursive:  #break if recursive is not set
                break

    elif option == "2":
        for cwd, subList, file in os.walk(
                path):  #start from selected directory and branch out
            if os.path.basename(cwd) != os.path.basename(path) and len(
                    file) != 0:
                # create new sheet

                sheetName = str(os.path.basename(cwd).encode('utf-8'))
                sheetName = re.sub(r'[^\w]', ' ', sheetName)
                sheetName = '{:1.29}'.format(
                    sheetName)  #max hardcoded sheetname 31 we limit to 30
                hashBook.create_sheet(sheetName)
                hashSheet = sheetHeader(hashBook, option,
                                        sheetNum)  # reprep sheet

                sheetNum += 1

            for name in file:

                i += 1
                print(
                    ">>> Processed: {} out of {} files".format(i, totalFiles),
                    end='\r',
                    flush=True)

                try:
                    with open((os.path.join(cwd, name)),
                              'rb') as hashMe:  #read in binary
                        contents = hashMe.read()  #read in the file
                    hashes = [name, md5File(contents)]
                    hashSheet.append(hashes)
                except IOError:
                    filesSkipped.append(os.path.join(cwd, name))
            if not recursive:  #break if recursive is not set
                break
    elif option == "3":
        for cwd, subList, file in os.walk(
                path):  #start from selected directory and branch out
            if os.path.basename(cwd) != os.path.basename(path) and len(
                    file) != 0:
                # create new sheet
                sheetName = str(os.path.basename(cwd).encode('utf-8'))
                sheetName = re.sub(r'[^\w]', ' ', sheetName)
                sheetName = '{:1.29}'.format(
                    sheetName)  #max hardcoded sheetname 31 we limit to 30
                hashBook.create_sheet(sheetName)
                hashSheet = sheetHeader(hashBook, option,
                                        sheetNum)  # reprep sheet

                sheetNum += 1
            for name in file:

                i += 1
                print(
                    ">>> Processed: {} out of {} files".format(i, totalFiles),
                    end='\r',
                    flush=True)

                try:
                    with open((os.path.join(cwd, name)),
                              'rb') as hashMe:  #read in binary
                        contents = hashMe.read()  #read in the file

                        hashes = [name, sh1File(contents)]
                        hashSheet.append(hashes)
                except IOError:
                    filesSkipped.append(os.path.join(cwd, name))

            if not recursive:  #break if recursive is not set
                break
    elif option == "4":
        for cwd, subList, file in os.walk(
                path):  #start from selected directory and branch out
            if os.path.basename(cwd) != os.path.basename(path) and len(
                    file) != 0:
                # create new sheet
                sheetName = str(os.path.basename(cwd).encode('utf-8'))
                sheetName = re.sub(r'[^\w]', ' ', sheetName)
                sheetName = '{:1.29}'.format(
                    sheetName)  #max hardcoded sheetname 31 we limit to 30
                hashBook.create_sheet(sheetName)
                hashSheet = sheetHeader(hashBook, option,
                                        sheetNum)  # reprep sheet

                sheetNum += 1
            for name in file:

                i += 1
                print(
                    ">>> Processed: {} out of {} files".format(i, totalFiles),
                    end='\r',
                    flush=True)
                try:
                    with open((os.path.join(cwd, name)),
                              'rb') as hashMe:  #read in binary
                        contents = hashMe.read()  #read in the file

                        hashes = [name, sh56File(contents)]
                        hashSheet.append(hashes)
                except IOError:
                    filesSkipped.append(os.path.join(cwd, name))

            if not recursive:  #break if recursive is not set
                break
    end = time.time()
    print(">>> Processed: {} out of {} files".format(
        i, totalFiles))  #reprint process to screen
    print("Completed in {} seconds".format(end - start))  #print time
    print("\n")
    if len(filesSkipped) > 0:
        print("Files skipped do to access errors")
        print("---------------------------------------")
        print(filesSkipped)
        print("\n")

    while True:
        try:  #save spreadsheet
            hashBook.save((os.path.join(output, filename)))
            print("File saved as: ",
                  os.path.join(os.path.abspath(output), filename))
            break  #break loop if save successful
        except PermissionError:  #catch and offer to save file elsewhere
            print(PermissionError)
            print("Unable to save spreadsheet, possibly file in use?")
            filename = input("Enter new file name to save as: ")
            filename += ".xlsx"


#md5 hash function, return digest
def md5File(contents):
    m5 = hashlib.md5()
    m5.update(contents)

    return m5.hexdigest()


#sha1 hash function, return digest
def sh1File(contents):
    sh1 = hashlib.sha1()
    sh1.update(contents)

    return sh1.hexdigest()


#sha256 hash function, return digest
def sh56File(contents):
    sh56 = hashlib.sha256()
    sh56.update(contents)

    return sh56.hexdigest()


# verify provided path or get path from user
def getPath(text, path=None):
    # if userFile defaults to None ask for file name
    if not path:
        path = input(text)

    while True:  #check path
        path = Path(path)  #set path to work for current os
        if str(path) == ".":  #current directory
            return os.getcwd()
        elif os.path.exists(path):  #return a valid path
            return path
        else:
            print("You entered: ", path, " which does not seem to exist")
            path = input("enter new path or 0 to exit: ")

            if path == "0":
                sys.exit(0)


#flexible function to return a valid user input or exit system
def validInput(option, first, last):
    while True:
        try:  #need to convert string to int and do not want any crashes here
            # check for quick exit
            if int(option) == 0:
                sys.exit(0)
            elif int(option) >= first and int(option) <= last:
                return option  #return valid option
            elif int(option) == 6:
                help()
                option = input("\nEnter new option or 0 to exit: ")
            else:
                option = input(
                    "Invalid option, please try again or enter 0 to exit: ")
        except ValueError:
            print(ValueError,
                  " now look what you've done, almost crashed the program!")
            option = input(
                "Try entering a number this time and remember 0 to exit: ")


def help():  #really more of an about page but same thing tbh
    #Print out some helpful information about program usage, and checksums
    print(
        "", 60 * '-', "\n About hash sheet generator", "\n", 60 * '-',
        "\n This program is designed to help you generate various",
        "\n checksums of files on your system. By selecting a hashing",
        "\n algorithm the program will ask for the directory name to check",
        "\n as well where to save your document (defaults to script directory)"
        "\n so that when calculations are finished you may open the",
        "\n document using excel or similar programs.")

    #show valid args
    validArgs()

    #remind user to backup files
    print("\n Reminder: always backup your files")


#display valid arguments and example
def validArgs():
    #print out a list of valid command line arguments
    print("\n\n Valid command line arguments:", "\n", 33 * '-',
          "\n First argument:", "\n", 16 * '-', "\n 1 -> MD5, SHA-1, & SHA-256",
          "\n 2 -> MD5", "\n 3 -> SHA-1", "\n 4 -> SHA-256",
          "\n 5 -> Single file mode", "\n 6 -> Help", "\n\n Second argument:"
          "\n", 16 * '-', "\n Full path", "\n Absolute path",
          "\n * -> Current path", "\n\n Third argument:", "\n", 16 * '-',
          "\n Full path", "\n Absolute path -> Includes current",
          "\n\n Fourth argument:"
          "\n", 16 * '-', "\n File name", "\n\n Fifth argument:", "\n",
          16 * '-', "\n r or y -> Recursive (optional)"
          "\n", 33 * '-')

    #example usage
    print("\n Example:" "\n", 9 * "-")
    print(
        " Relative: python filehash.py 1 Pictures outputdirectory outputfile r")
    print(
        " Absolute: python filehash.py 1 C:\\Program Files C:\\Users\\example\\Documents\\ outputfile r"
    )
    print(" Current: python filehash.py 1 . outputdirectory outputfile y")


#call to main at the end
if __name__ == '__main__':
    totalFiles = 0  #a global variable for total file count
    main()
